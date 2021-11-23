from openpyxl import load_workbook

workbook = load_workbook('./avtovoz_cars.xlsx')
sheet = workbook['Лист1']

for row in sheet.rows:
    for cell in row:
        if cell.value == 'Audi':
            print(cell.value, sheet.cell(row=cell.row, column=2).value)
